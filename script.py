from docx import Document, document
import openpyxl
import os

# Paths
template_path = "./ChangeMe.docx"
output_dir = "factures"
os.makedirs(output_dir, exist_ok=True)
excel_file = "word_metadata.xlsx"

# Create or load Excel workbook
if os.path.exists(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["File Name", "Content Summary"])

# Sample data to populate the table
data = [
    {"khedma": "faza", "qte": "2", "Prix": "400"},
    {"khedma": "faza2", "qte": "1", "Prix": "400"},
    {"khedma": "faza3", "qte": "1", "Prix": "400"},
    
]

# Generate Word files using the template with a table
for i, record in enumerate(data, start=1):
    # Load the Word template
    doc = Document(template_path)
    
    # Access the first table in the document (adjust index if there are multiple tables)
    table = doc.tables[1]
    
    # Populate the table
    for row_index, (key, value) in enumerate(record.items(), start=1):
        # Populate each row in the table (assuming row 0 is the header)
        table.cell(row_index, 1).text = record  # khedma
        table.cell(row_index, 2).text = value  # qte
        table.cell(row_index, 3).text = "100" # prix unitaire

    # Save the new Word file
    file_name = f"Document_{i}.docx"
    file_path = os.path.join(output_dir, file_name)
    doc.save(file_path)

    # Add metadata to Excel
    ws.append([file_name, f"Record for {record['khedma']}"])  # Save summary

# Save Excel file
wb.save(excel_file)

print("Word files with table data created, and metadata added to Excel.")
